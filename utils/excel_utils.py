import openpyxl

from config.config import EXCEL_FILE


def read_excel():
    # 打开 excel 文件
    # workbook = openpyxl.load_workbook(file_path)
    workbook = openpyxl.load_workbook(EXCEL_FILE)

    # 读取数据操作
    #思路：先把所有用例数据读出来，再进行筛选，合格的用例添加到data中
    all_cases = []
    #定义当前用例数据，处理每一条
    current_case = []

    data = []  # 空列表，用于组装字典

    #加个循环
    for worksheet in workbook.worksheets:

        keys = [cell.value for cell in worksheet[2]]

        for row in worksheet.iter_rows(min_row=3, values_only=True): # 从第三行开始逐行读取，只返回值
            dict_data = dict(zip(keys, row))
            #如果读取的is_true是TRUE则append，否则，不append
            # if dict_data["is_true"]:
            #     data.append(dict_data)

            #问题1 合并单元格出现空值
            #解决思路1  id不为none，就视为用例
            #          只要id为none，就视为步骤，只要中间两行'step_num'和 'step_name'，确保当前有用例数据的情况下把步骤添加进去
            if dict_data["id"] is not None:
                #组织用例的过程
                current_case = {
                    "id": dict_data["id"],
                    "feature": dict_data["feature"],
                    "story": dict_data["story"],
                    "title": dict_data["title"],
                    "steps": [{ #"steps" 设计成 列表，列表里的每个元素用 字典 表示“一条步骤”的字段
                        "step_num": dict_data["step_num"],
                        "step_name": dict_data["step_name"],
                        "keyword": dict_data["keyword"],
                        "by": dict_data["by"],
                        "value": dict_data["value"],
                        "data": dict_data["data"],
                        "index": dict_data["index"],
                    }],
                    "is_true": dict_data["is_true"],
                }
                all_cases.append(current_case) #临时存到all_cases = []


            elif current_case is not None:
                current_case["steps"].append({
                    "step_num": dict_data["step_num"],
                    "step_name": dict_data["step_name"],
                    "keyword": dict_data["keyword"],
                    "by": dict_data["by"],
                    "value": dict_data["value"],
                    "data": dict_data["data"],
                    "index": dict_data["index"],
                })

    #过滤用例数据，只保留is_true为true的用例
    data = [case for case in all_cases if case["is_true"]]
    print("**  data   **", data)  # 打印拿到的所有数据

    #问题2 用例条数不对，要把内部步骤合并









    # 关闭 excel 文件
    workbook.close()

    return data

# read_excel()